import streamlit as st
import pandas as pd
import re
import io
from extractor import process_file
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, numbers as xl_numbers

st.set_page_config(page_title="Bank Statement Analyser", layout="wide")
st.title("🏦 Bank Statement Analyser")

uploaded_file = st.file_uploader("Upload Bank Statement (PDF)", type=["pdf"])


# ── DATE CLEANING ────────────────────────────────────────────────────────────
def clean_date(value):
    if not value:
        return None
    text = str(value).strip()
    match = re.search(r"(\d{1,2})[\/\-.](\d{1,2})[\/\-.](\d{2,4})", text)
    if match:
        d, m, y = match.groups()
        if len(y) == 2:
            y = "20" + y
        return f"{int(d):02d}/{int(m):02d}/{y}"
    return text  # return original if no pattern match


# ── AMOUNT CLEANING ──────────────────────────────────────────────────────────
def clean_amount(value):
    if value is None or str(value).strip() in ("", "None", "nan"):
        return 0.00
    try:
        return round(float(str(value).replace(",", "").strip()), 2)
    except Exception:
        return 0.00


# ═══════════════════════════════════════════════════════════════════════════
#                               MAIN PROCESS
# ═══════════════════════════════════════════════════════════════════════════
if uploaded_file is not None:

    st.info(f"⏳ Processing: **{uploaded_file.name}** ...")
    file_bytes = uploaded_file.read()

    try:
        meta, transactions = process_file(file_bytes, uploaded_file.name)
    except Exception as e:
        st.error(f"Extraction failed: {e}")
        st.stop()

    if not transactions:
        st.error("⚠️ No transactions found. Check PDF format.")
        st.stop()

    df = pd.DataFrame(transactions)

    # ── 1. CLEAN ─────────────────────────────────────────────────────────────
    df_final = df.copy()

    if "Date" in df_final.columns:
        df_final["Date"] = df_final["Date"].apply(clean_date)

    for col in ["Debit", "Credit", "Balance"]:
        if col in df_final.columns:
            df_final[col] = df_final[col].apply(clean_amount)

    # Drop internal Page column if present
    if "Page" in df_final.columns:
        df_final.drop(columns=["Page"], inplace=True)

    # Reorder: Date, Particulars, Debit, Credit, Balance, Head
    preferred_order = ["Date", "Particulars", "Debit", "Credit", "Balance", "Head"]
    existing_cols = [c for c in preferred_order if c in df_final.columns]
    extra_cols = [c for c in df_final.columns if c not in existing_cols]
    df_final = df_final[existing_cols + extra_cols]

    # ── 2. SUMMARY METRICS ───────────────────────────────────────────────────
    total_debit  = df_final["Debit"].sum()  if "Debit"  in df_final.columns else 0
    total_credit = df_final["Credit"].sum() if "Credit" in df_final.columns else 0
    row_count    = len(df_final)

    c1, c2, c3 = st.columns(3)
    c1.metric("Total Transactions", row_count)
    c2.metric("Total Debit",  f"₹ {total_debit:,.2f}")
    c3.metric("Total Credit", f"₹ {total_credit:,.2f}")

    # ── 3. DISPLAY ───────────────────────────────────────────────────────────
    df_display = df_final.copy()
    for col in ["Debit", "Credit", "Balance"]:
        if col in df_display.columns:
            df_display[col] = df_display[col].map(lambda x: f"{x:,.2f}")

    st.success(f"✅ {row_count} transactions extracted!")
    st.dataframe(df_display, use_container_width=True)

    # ── Base names ────────────────────────────────────────────────────────────
    base      = uploaded_file.name.rsplit(".", 1)[0]
    csv_name  = f"{base}.csv"
    xlsx_name = f"{base}.xlsx"
    pdf_name  = f"{base}_report.pdf"

    # ── 4. CSV ───────────────────────────────────────────────────────────────
    csv_bytes = df_final.to_csv(index=False, float_format="%.2f").encode("utf-8")

    # ── 5. EXCEL ─────────────────────────────────────────────────────────────
    excel_buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    HEADER_FILL = PatternFill("solid", fgColor="2F5496")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    NUM_COLS    = {"Debit", "Credit", "Balance"}

    for r_idx, row_data in enumerate(dataframe_to_rows(df_final, index=False, header=True), start=1):
        ws.append(row_data)
        for c_idx, cell in enumerate(ws[r_idx], start=1):
            header_val = ws.cell(1, c_idx).value or ""
            if r_idx == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                if header_val in NUM_COLS:
                    cell.number_format = xl_numbers.FORMAT_NUMBER_00
                    cell.alignment = Alignment(horizontal="left")
                else:
                    cell.alignment = Alignment(horizontal="left", wrap_text=True)

    # Auto column width
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 50)

    ws.freeze_panes = "A2"
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # ── 6. PDF ───────────────────────────────────────────────────────────────
    df_pdf = df_final.copy()
    for col in ["Debit", "Credit", "Balance"]:
        if col in df_pdf.columns:
            df_pdf[col] = df_pdf[col].map(lambda x: f"{x:,.2f}")

    pdf_buffer  = io.BytesIO()
    styles      = getSampleStyleSheet()
    doc         = SimpleDocTemplate(
        pdf_buffer, pagesize=landscape(A4),
        leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20
    )
    elements = []

    # Title
    elements.append(Paragraph(f"Bank Statement — {base}", styles["Title"]))
    elements.append(Spacer(1, 10))

    data = [list(df_pdf.columns)] + [
        [str(v) for v in row] for row in df_pdf.values.tolist()
    ]

    col_count = len(df_pdf.columns)
    col_widths = []
    for c in df_pdf.columns:
        if c == "Particulars":
            col_widths.append(220)
        elif c == "Date":
            col_widths.append(65)
        else:
            col_widths.append(70)

    table = Table(data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        # Header
        ("BACKGROUND",   (0, 0), (-1, 0),  colors.HexColor("#2F5496")),
        ("TEXTCOLOR",    (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, 0),  8),
        ("ALIGN",        (0, 0), (-1, 0),  "CENTER"),
        ("VALIGN",       (0, 0), (-1, 0),  "MIDDLE"),
        # Body
        ("FONTNAME",     (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",     (0, 1), (-1, -1), 7),
        ("ALIGN",        (0, 1), (-1, -1), "LEFT"),
        ("VALIGN",       (0, 1), (-1, -1), "MIDDLE"),
        # Alternating rows
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF2F8")]),
        # Grid
        ("GRID",         (0, 0), (-1, -1), 0.25, colors.grey),
        ("TOPPADDING",   (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
        ("LEFTPADDING",  (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))

    elements.append(table)
    doc.build(elements)
    pdf_buffer.seek(0)

    # ── 7. DOWNLOADS ─────────────────────────────────────────────────────────
    st.markdown("---")
    dl1, dl2, dl3 = st.columns(3)

    with dl1:
        st.download_button("⬇️ Download CSV",
                           data=csv_bytes,
                           file_name=csv_name,
                           mime="text/csv")
    with dl2:
        st.download_button("⬇️ Download Excel",
                           data=excel_buffer,
                           file_name=xlsx_name,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with dl3:
        st.download_button("⬇️ Download PDF",
                           data=pdf_buffer,
                           file_name=pdf_name,
                           mime="application/pdf")
